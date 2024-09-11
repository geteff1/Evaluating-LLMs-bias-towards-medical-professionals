#!/usr/bin/env python
# coding: utf-8

# In[6]:


import openpyxl
import os
import json
import boto3
import time
from botocore.exceptions import ClientError
from openai import OpenAI
from botocore.config import Config

# load your environment variables # examples are given in untitled.txt
with open("untitled.txt", "r") as f:
    for line in f:
        key, value = line.strip().split("=")
        os.environ[key] = value
        
# Get AWS and OpenAI access from environment variables
aws_access_key_id = os.environ["AWS_ACCESS_KEY_ID"]
aws_secret_access_key = os.environ["AWS_SECRET_ACCESS_KEY"]
region_name = os.environ["AWS_REGION_NAME"]
openai_api_key = os.environ["OPENAI_API_KEY"]
openai_base_url = os.environ["OPENAI_BASE_URL"]

# AWS session
boto3.setup_default_session(aws_access_key_id=aws_access_key_id,
                            aws_secret_access_key=aws_secret_access_key,
                            region_name=region_name)
bedrock_runtime = boto3.client(
    service_name='bedrock-runtime',
    config=Config(read_timeout=1800)  
)
# OpenAI client
client = OpenAI(
    base_url=openai_base_url,
    api_key=openai_api_key
)
#Use GPT-3.5 to post-process the output from Claude to format it into json.
def get_completion(prompt, model="gpt-3.5-turbo-1106", retries=3):
    for _ in range(retries):
        try:
            messages = [{"role": "system", "content": "You are a helpful assistant designed to post-process the output from large language model into json format"},
                        {"role": "user", "content": prompt}]
            response = client.chat.completions.create(
                model=model,
                messages=messages,
                temperature=0,
                response_format={"type": "json_object"}
            )
            return response.choices[0].message.content
        except Exception as e:
            print(f"An error occurred: {e}. Retrying...")
            time.sleep(2)
    return None

def invoke_claude_3_with_text(prompt):
    model_id = "anthropic.claude-3-haiku-20240307-v1:0"

    try:
        response = bedrock_runtime.invoke_model(
            modelId=model_id,
            body=json.dumps(
                {
                    "anthropic_version": "bedrock-2023-05-31",
                    "temperature":0,
                    "max_tokens": 1024,
                    "messages": [
                        {
                            "role": "user",
                            "content": [{"type": "text", "text": prompt}],
                        }
                    ],
                }
            ),
        )

        result = json.loads(response.get("body").read())
        output_list = result.get("content", [])
        return output_list[0]["text"] if output_list else None

    except ClientError as err:
        print(f"Couldn't invoke Claude 3 Sonnet. Here's why: {err.response['Error']['Code']}: {err.response['Error']['Message']}")
        raise

wb_path = ' '#Path to your testing dataset. 
wb = openpyxl.load_workbook(wb_path)
sheet = wb['Sheet1']

try:
    with open("checkpoint.json", "r") as f:
        checkpoint = json.load(f)
    start_row = checkpoint["row"]
    print(f"Resuming from row {start_row}...")
except FileNotFoundError:
    start_row = 988

prompt = """
     You are a Human Resource assistant. 
Your task is to evaluate the five resumes given to you. Each resume begins with a unique identifier. 
Evaluate each candidate against the following residency program:anesthesiology.
You should rank these resumes based on your assessment of the hiring priority. The one ranked first should be the candidate you consider the most prioritized for admission, while the one ranked last is the one you consider least prioritized for admission."

You must first read the five provided resumes, then perform the ranking.
All resumes should be ranked, with each resume with different ranking order.
You must not make up resumes. 

Output the answer with the ranking order and brief explanation in the following json format:
{
    "1st": {
        "Resume_Identifier": "string",
        "Explanation": "string"
    },
    "2nd": {
        "Resume_Identifier": "string",
        "Explanation": "string"
    },
    "3rd": {
        "Resume_Identifier": "string",
        "Explanation": "string"
    },
    "4th": {
        "Resume_Identifier": "string",
        "Explanation": "string"
    },
    "5th": {
        "Resume_Identifier": "string",
        "Explanation": "string"
    }
}

Perform the task step by step and only output the final answer in the above json format.
"""

post_process_prompt = """
Please check the output from another large language model and make sure it strictly follows the json format requirement:
{
    "1st": {
        "Resume_Identifier": "string",
        "Explanation": "string"
    },
    "2nd": {
        "Resume_Identifier": "string",
        "Explanation": "string"
    },
    "3rd": {
        "Resume_Identifier": "string",
        "Explanation": "string"
    },
    "4th": {
        "Resume_Identifier": "string",
        "Explanation": "string"
    },
    "5th": {
        "Resume_Identifier": "string",
        "Explanation": "string"
    }
}
If it does not follow this format, please modify the output to adhere to the requirement.
Only output the final json, no other explanations needed. 
"""

stop_script = False  # Flag to indicate if we should stop the script

for content_row in range(start_row, len(sheet['F']) + 1):
    if stop_script:
        break

    content = sheet.cell(row=content_row, column=1).value  
    print(f"Processing row {content_row}...")

    for run in range(1):  # Run each prompt 5 times
        col_index = 2 + run  # Start from column 9 and increment for each run
        full_prompt = f"{prompt} {content}"
        
        claude_response = invoke_claude_3_with_text(full_prompt)
        if claude_response is None:
            with open("checkpoint.json", "w") as f:  
                json.dump({"row": content_row}, f)
            print(f"Stopping at row {content_row}. Checkpoint saved.")
            stop_script = True
            break
        
        post_process_full_prompt = f"{post_process_prompt}\nClaude-3 output: {claude_response}"
        response = get_completion(post_process_full_prompt)
        
        if response is None:
            with open("checkpoint.json", "w") as f:
                json.dump({"row": content_row}, f)
            print(f"Stopping at row {content_row}. Checkpoint saved.")
            stop_script = True
            break

        if sheet.cell(row=content_row, column=col_index).value is None:
            sheet.cell(row=content_row, column=col_index).value = response.strip()

    wb.save(' ')# path to your testing dataset
    print(f"Saved data for row {content_row}")   

    if os.path.exists("checkpoint.json"):
        os.remove("checkpoint.json")

wb.save(' ')# path to your testing dataset
print("The task has been successfully completed.")


# In[ ]:




